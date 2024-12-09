import openpyxl

class Report:
    def __init__(self, file_path):
        self.file = file_path
        self.wb = openpyxl.load_workbook(self.file)
        self.sheet = [self.wb[sheet_name] for sheet_name in self.wb.sheetnames]
        self.csv = [self.sheet_to_csv(sheet) for sheet in self.sheet]

    def __str__(self):
        return f"Report({self.file})"

    def __repr__(self):
        return f"Report({self.file})"
    
    def get_report(self):
        result = ''
        for csv in self.csv:
            result += csv + '\n'
        return result

    def sheet_to_csv(self, sheet):
        result = ''
        for row in sheet.iter_rows():
            for cell in row:
                result += str(cell.value) + ','
            result = result[:-1] + '\n'
        
        return result
